import csv
import re
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import List, Callable, Dict, Iterable, TextIO, Any, Sequence
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
import pandas as pd
import yaml
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from simpler_core.plugin import DataSourcePlugin, DataSourceType, EntityLink
from simpler_core.schema import load_external_schema_from_yaml
from simpler_model import Entity, Relation, Attribute
from simpler_plugin_json import JSONDataSourcePlugin


class TabularDataSourceType(DataSourceType):
    name = 'Tabular'
    inputs = [
        'data_no_header',  # A zip file containing all partial CSV files
        'data_header',     # A zip file containing all partial CSV files
        'schema'
    ]
    input_validation_statement = r'(data_header.*|data_no_header.*)'


def extract_csv_data(directory: Path, reader_factory: Callable[[TextIO], Iterable]) -> Dict:
    result = {}
    for file in directory.glob('*.csv'):
        with open(file, 'r') as stream:
            reader = reader_factory(stream)
            result[file.stem] = list(reader)
    return result


class TabularDataSourcePlugin(DataSourcePlugin):

    data_source_type = TabularDataSourceType()

    def get_strong_entities(self, name: str) -> List[Entity]:
        pass

    def get_all_entities(self, name: str) -> List[Entity]:
        no_header_data = None
        with_header_data = None
        schema = None
        with self.storage.get_data(name) as stream_lookup, \
                TemporaryDirectory() as zip_no_header_directory, \
                TemporaryDirectory() as zip_directory:
            if 'data_no_header' in stream_lookup:
                with ZipFile(stream_lookup['data_no_header']) as zip_handle:
                    zip_handle.extractall(zip_no_header_directory)
                no_header_data = extract_csv_data(Path(zip_no_header_directory), csv.reader)
            if 'data_header' in stream_lookup:
                with ZipFile(stream_lookup['data_header']) as zip_handle:
                    zip_handle.extractall(zip_directory)
                with_header_data = extract_csv_data(Path(zip_directory), lambda stream: csv.DictReader(stream))
            if 'schema' in stream_lookup:
                schema_data = load_external_schema_from_yaml(stream_lookup['schema'])
                schema = {
                    entity.entity_name[0]: entity
                    for entity_data in schema_data
                    for entity in [Entity.from_dict(entity_data)]
                }

        if with_header_data is None:
            with_header_data = {}
        for key, table in no_header_data.items():
            with_header_data[key] = [
                {
                    f'Column{idx}': cell
                    for idx, cell in enumerate(row)
                }
                for row in table
            ]

        entities = []
        for entity_name, entity_data in with_header_data.items():
            if schema is not None and entity_name in schema:
                entities.append(schema[entity_name])
            elif schema is None:
                entity = Entity(
                    has_attribute=[],
                    has_entity_modifier=[],
                    is_object_in_relation=[],
                    is_subject_in_relation=[],
                    entity_name=[entity_name]
                )
                attributes = []
                for heading in entity_data[0].keys():
                    attribute = Attribute(
                        attribute_name=[heading],
                        is_attribute_of=None,
                        has_attribute_modifier=[]
                    )
                    attributes.append(attribute)
                entity.has_attribute = attributes
                entities.append(entity)

        return entities

    def get_related_entity_links(self, name: str) -> List[Relation]:
        pass

    def get_entity_by_id(self, name: str, entity_id: str) -> Entity:
        pass


def get_table_ref_without_header(table: Table) -> str:
    start, end = table.ref.split(':')
    column, row = re.match(r'([A-Z]*)(\d+)', start).groups()
    new_row = str(int(row) + table.headerRowCount)
    return f'{column}{new_row}:{end}'


def get_index_range_of_ref_string(ref: str) -> range:
    start, end = ref.split(':')
    _, start_row = re.match(r'([A-Z]*)(\d+)', start).groups()
    _, end_row = re.match(r'([A-Z]*)(\d+)', end).groups()

    return range(int(start_row), int(end_row) + 1)


# powered by MS copilot
def column_string_to_number(column_string):
    column_string = column_string.upper()
    column_number = 0
    for char in column_string:
        column_number = column_number * 26 + (ord(char) - ord('A') + 1)
    return column_number - 1


def get_column_count_from_data_ref(ref: str) -> int:
    start, end = ref.split(':')
    start_column, _ = re.match(r'([A-Z]+)(\d*)', start).groups()
    end_column, _ = re.match(r'([A-Z]+)(\d*)', end).groups()
    start_number = column_string_to_number(start_column)
    end_number = column_string_to_number(end_column)
    return end_number - start_number + 1


def data_generator(sheet: Worksheet, ref: str):
    start, end = ref.split(':')
    start_column, start_row = re.match(r'([A-Z]+)(\d+)', start).groups()
    end_column, end_row = re.match(r'([A-Z]+)(\d+)', end).groups()

    start_row_number = int(start_row)
    end_row_number = int(end_row)

    start_column_number = column_string_to_number(start_column) + 1
    end_column_number = column_string_to_number(end_column) + 1

    for row in sheet.iter_rows(
        min_row=start_row_number,
        max_row=end_row_number,
        min_col=start_column_number,
        max_col=end_column_number
    ):
        yield [cell.value for cell in row]


@dataclass
class ExcelTableDefinition:
    index: pd.MultiIndex | pd.Index | List[int | str] | range
    columns: pd.MultiIndex | pd.Index | List[str]
    data_ref: str
    worksheet: Worksheet
    table_name: str

    def get_frame(self) -> pd.DataFrame:
        return pd.DataFrame(
            data=data_generator(self.worksheet, self.data_ref),
            index=self.index,
            columns=self.columns
        )

    @staticmethod
    def from_native_table(table: Table, worksheet: Worksheet) -> 'ExcelTableDefinition':
        data_ref = get_table_ref_without_header(table)
        return ExcelTableDefinition(
            index=get_index_range_of_ref_string(data_ref),
            columns=table.column_names,
            data_ref=data_ref,
            worksheet=worksheet,
            table_name=table.name
        )

    @staticmethod
    def from_custom_table_def(
            worksheet: Worksheet,
            header_ref: str | None,
            data_ref: str,
            name: str
    ) -> 'ExcelTableDefinition':
        if header_ref is not None:
            column_data = list(data_generator(worksheet, header_ref))
            if len(column_data) == 1:
                columns = column_data[0]
            else:
                columns = pd.MultiIndex.from_arrays(column_data)
        else:
            columns = [
                f'Column{i}'
                for i in range(1, get_column_count_from_data_ref(data_ref) + 1)
            ]
        return ExcelTableDefinition(
            index=get_index_range_of_ref_string(data_ref),
            columns=columns,
            data_ref=data_ref,
            worksheet=worksheet,
            table_name=name
        )


class ExcelDataSourceType(DataSourceType):
    name = 'Excel'
    inputs = [
        'workbook.xlsx',  # The XLSX Document
        'table_def.yaml'
    ]
    input_validation_statement = r'.*workbook\.xlsx'


class ExcelDataSourcePlugin(DataSourcePlugin):

    data_source_type = ExcelDataSourceType()

    def _get_tables(self, name: str) -> Dict[str, pd.DataFrame]:
        tables: Dict[str, pd.DataFrame] = {}
        table_definitions: List[ExcelTableDefinition] = []
        with self.storage.get_data(name) as stream_lookup:
            # workbook = load_workbook(stream_lookup['workbook.xlsx'], read_only=True)
            workbook = load_workbook(stream_lookup['workbook.xlsx'])
            # reader = pd.ExcelReader(stream_lookup['workbook.xlsx'])

            sheets = [workbook[name] for name in workbook.sheetnames]
            # native_tables = [table for sheet in sheets for table in sheet.tables.values()]
            table_definitions.extend([
                ExcelTableDefinition.from_native_table(
                    table,
                    sheet
                )
                for sheet in sheets
                for table in sheet.tables.values()
            ])

            if 'table_def.yaml' in stream_lookup:
                parsed_definitions = yaml.safe_load(stream_lookup['table_def.yaml'])
                table_definitions.extend([
                    ExcelTableDefinition.from_custom_table_def(
                        workbook[definition['sheet_name']],
                        definition['header_ref'],
                        definition['data_ref'],
                        definition['name']
                    )
                    for definition in parsed_definitions
                ])
            datetime_types = [
                'datetime64[ns]', # 'datetime64[ns,tz]',
                'datetime64[ms]', 'datetime64[s]',
                'datetime64[m]', 'datetime64[h]'
            ]
            for definition in table_definitions:
                frame = definition.get_frame()
                for col in frame.select_dtypes(include='datetime64').columns:
                    frame[col] = frame[col].apply(
                        lambda val: val.isoformat() if isinstance(val, pd.Timestamp) else val
                    ).astype('object').replace({pd.NaT: None})
                tables[definition.table_name] = frame

        return tables

    def get_strong_entities(self, name: str) -> List[Entity]:
        pass

    def get_all_entities(self, name: str) -> List[Entity]:
        tables = self._get_tables(name)
        obj = {
            key: frame.to_dict(orient='records')
            for key, frame in tables.items()
        }
        return JSONDataSourcePlugin.generate_model_from_dict(obj)

    def get_related_entity_links(self, name: str) -> List[EntityLink]:
        pass

    def get_entity_by_id(self, name: str, entity_id: str) -> Entity:
        pass
